import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference


def update_excel_summary(excel_path, bank):
    if not os.path.exists(excel_path):
        return

    df = pd.read_excel(excel_path)
    if df.empty:
        return

    # Add bank column if missing
    if "Bank" not in df.columns:
        df["Bank"] = bank

    # Clean up and standardize date field
    date_col = "Datetime" if "Datetime" in df.columns else "Date"
    df[date_col] = pd.to_datetime(df[date_col])
    df["Month"] = df[date_col].dt.to_period("M").astype(str)

    # Group by Month & Bank
    monthly_summary = df.groupby(["Month", "Bank"])["Amount"].sum().reset_index()

    # Load workbook
    wb = load_workbook(excel_path)

    # Remove old summary if exists
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    # Write DataFrame to sheet
    for r in dataframe_to_rows(monthly_summary, index=False, header=True):
        ws.append(r)

    # Add line chart for monthly spend
    chart = LineChart()
    chart.title = "Monthly Spend by Bank"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Amount"

    max_row = len(monthly_summary) + 1
    chart_data = Reference(ws, min_col=3, min_row=1, max_row=max_row)
    chart_cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    ws.add_chart(chart, "E2")

    wb.save(excel_path)
